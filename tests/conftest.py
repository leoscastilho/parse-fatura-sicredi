"""Fixtures compartilhadas.

Os testes são HERMÉTICOS: geram os próprios extratos em vez de depender de
`input/*.xls`, que é gitignored e não entra na imagem Docker. Sem isso a suíte
passaria na sua máquina e sumiria no build.

Quando os extratos reais existem, um teste extra roda contra eles — mas nenhum
teste obrigatório depende disso.
"""

from __future__ import annotations

import io
import shutil
from datetime import datetime
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook

REPO = Path(__file__).resolve().parent.parent
# A configuração REAL do app — a que o portal edita. Só um teste olha para ela,
# e apenas para conferir que carrega.
CONFIG_REAL = REPO / "config"
# A configuração dos TESTES: uma fotografia congelada. É contra ela que a suíte
# inteira roda, para que editar suas regras no portal nunca derrube o build.
CONFIG = Path(__file__).resolve().parent / "fixtures" / "config"
REAL_INPUT = REPO / "input"


# ---------------------------------------------------------------------------
# Extratos sintéticos
# ---------------------------------------------------------------------------

def _sicredi_workbook(path: Path, *, vencimento="10/08/2026", rows=None,
                      internacional=None) -> Path:
    """Monta uma planilha com o mesmo formato do extrato Sicredi."""
    rows = rows if rows is not None else [
        ("21/08/2024", "MERCADOLIVRE*ESPLAN", "08/10", "86,80"),
        ("26/06/2026", "Prudent*APOL00188798", None, "521,58"),
        ("02/07/2026", "SUPERMERCADOS ALVORA", None, "270,51"),
        ("03/07/2026", "OggiSantaRita", None, "32,70"),
        ("05/07/2026", "AUTO POSTO Z LTDA", None, "182,07"),
        ("07/07/2026", "AMAZON BR", None, "59,13"),
        ("08/07/2026", "LOJA XPTO", None, "42,00"),
        ("01/07/2026", "Pag Fat Deb Cc", None, "-13.928,90"),
        ("04/07/2026", "Credito Anuidade Car", None, "-75,00"),
    ]
    total = sum(
        float(v.replace(".", "").replace(",", ".")) for _, _, _, v in rows
        if float(v.replace(".", "").replace(",", ".")) > 0
    )
    creditos = -sum(
        float(v.replace(".", "").replace(",", ".")) for _, _, _, v in rows
        if float(v.replace(".", "").replace(",", ".")) < 0
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio"
    grid = [
        [None, None, None, None],
        ["Associado:", "TESTE", None, None],
        ["Cartão Sicredi MASTER - Extrato Mensal", None, None, None],
        ["Data de Vencimento:", None, vencimento, None],
        ["Valor Total(R$):", None, f"{total:,.2f}".replace(",", "@").replace(".", ",").replace("@", "."), None],
        ["Pagamentos / Créditos (R$):", None, f"{creditos:.2f}".replace(".", ","), None],
        [None, None, None, None],
        ["Histórico de Despesas", None, None, None],
        ["Despesas no Brasil", None, None, None],
        ["Data", "Descrição", "Parcela", "Valor (R$)"],
    ]
    grid += [[d, desc, parc, val] for d, desc, parc, val in rows]
    grid.append(["Valor Total R$:", None, None,
                 f"{total:,.2f}".replace(",", "@").replace(".", ",").replace("@", ".")])
    grid.append([None, None, None, None])
    grid.append(["Despesas Internacionais", None, None, None])
    grid.append(["Data", "Descrição", "Valor (US$)", "Valor (R$)"])
    if internacional:
        for d, desc, usd, brl in internacional:
            grid.append([d, desc, usd, brl])
        grid.append(["Valor Total R$:", None, None, internacional[-1][3]])
    else:
        grid.append(["Não existem lançamentos.", None, None, None])

    for row in grid:
        ws.append(row)
    wb.save(path)
    return path


def _sicredi_xls(path: Path, **kwargs) -> Path:
    """O MESMO extrato, gravado como `.xls` de verdade — BIFF, não zip.

    Existe por causa do CONTAINER, não do conteúdo: o `.xls` antigo mora no
    mesmo OLE2 de um `.xlsx` cifrado, e é o único arquivo do projeto capaz de
    provar que o portal não confunde os dois. Com `.xlsx` em toda fixture, "isto
    está protegido por senha?" nunca era perguntado a um OLE2 aberto — e a
    resposta errada faria o portal pedir senha para uma fatura do Sicredi.

    De quebra é o único teste que exercita o caminho do `xlrd`, que está nos
    requisitos justamente para este formato.
    """
    import xlwt

    aberto = _sicredi_workbook(path.with_suffix(".xlsx"), **kwargs)
    quadro = pd.read_excel(aberto, sheet_name=0, header=None, dtype=object)

    wb = xlwt.Workbook()
    ws = wb.add_sheet("Relatorio")
    for r, linha in enumerate(quadro.itertuples(index=False)):
        for c, valor in enumerate(linha):
            if valor is not None and not (isinstance(valor, float) and pd.isna(valor)):
                ws.write(r, c, str(valor))
    wb.save(str(path))
    return path


@pytest.fixture
def sicredi_xlsx(tmp_path) -> Path:
    return _sicredi_workbook(tmp_path / "extrato-sicredi.xlsx")


@pytest.fixture
def sicredi_xls(tmp_path) -> Path:
    return _sicredi_xls(tmp_path / "extrato-sicredi.xls")


@pytest.fixture
def sicredi_xlsx_intl(tmp_path) -> Path:
    """Extrato com seção internacional — a que o parser antigo perdia."""
    return _sicredi_workbook(
        tmp_path / "extrato-intl.xlsx",
        internacional=[("27/06/2026", "CLOUDFLARE", "20,92", "109,23")],
    )


def _sicredi_app_csv(path: Path, *, vencimento="10/09/2025", rows=None,
                     nomes=None) -> Path:
    """O CSV que o APLICATIVO do Sicredi exporta — o outro formato do mesmo banco.

    Reproduz as três coisas que o separam do `.xls` do site e que o leitor
    precisa vencer: o preâmbulo de rótulos antes da tabela, o `;` como
    separador, e a parcela escrita entre parênteses.
    """
    rows = rows if rows is not None else [
        ("20/08/2025", "MERCADOLIVRE UNICAF", "(01/02)", "R$ 150,00", ""),
        ("14/08/2025", "CE PERU RAIL 1", "", "R$ 320,59", "U$ 58,25"),
        ("02/07/2025", "SUPERMERCADOS ALVORA", "", "R$ 270,51", ""),
        ("01/07/2025", "Pagamento Efetuado", "", "R$ -400,00", ""),
    ]
    brasil = sum(_valor(v) for _, _, _, v, dolar in rows if _valor(v) > 0 and not dolar)
    exterior = sum(_valor(v) for _, _, _, v, dolar in rows if _valor(v) > 0 and dolar)
    creditos = sum(_valor(v) for _, _, _, v, _ in rows if _valor(v) < 0)

    def brl(v):
        return f'"R$ {v:,.2f}"'.replace(",", "\x00").replace(".", ",").replace("\x00", ".")

    linhas = [
        f" Associado ;{(nomes or ['Fulano de Tal'])[0]};;;;",
        " Cooperativa ;0230;;;;",
        "",
        f" Data de Vencimento ;{vencimento};;;;",
        f" Valor Total (R$) ;{brl(brasil + exterior + creditos)};;;;",
        " Situação ;Fechada;;;;",
        "",
        " Resumo de Despesas ;;;;;",
        f" (-) Pagamentos / Creditos (R$) ;{brl(creditos)};;;;",
        f" (+) Despesas / Debitos no Brasil (R$) ;{brl(brasil)};;;;",
        f" (+) Despesas / Debitos no exterior (R$) ;{brl(exterior)};;;;",
        f" (=) Total desta fatura (R$) ;{brl(brasil + exterior + creditos)};;;;",
        "",
        " Data ; Descrição ; Parcela ; Valor ; Valor em Dólar ; Adicional ; Nome;",
    ]
    # Conta conjunta: `nomes` dá um titular por linha, ciclicamente. Sem ele,
    # todas as compras são do mesmo dono — o cartão de uma pessoa só.
    titulares = nomes or ["Fulano de Tal"]
    for i, (data, desc, parcela, valor, dolar) in enumerate(rows):
        quem = titulares[i % len(titulares)]
        linhas.append(f'{data};{desc};{parcela};"{valor}";{dolar};;{quem}')
    # O app grava com BOM; `utf-8-sig` na escrita é o que o reproduz.
    path.write_text("\n".join(linhas) + "\n", encoding="utf-8-sig")
    return path


def _valor(texto: str) -> float:
    import re as _re
    limpo = _re.sub(r"[^\d,.\-]", "", texto).replace(".", "").replace(",", ".")
    return float(limpo) if limpo else 0.0


@pytest.fixture
def sicredi_app_csv(tmp_path) -> Path:
    return _sicredi_app_csv(tmp_path / "fatura-app.csv")


# ---------------------------------------------------------------------------
# BTG Pactual — planilha em tabelas empilhadas, e cifrada
# ---------------------------------------------------------------------------

SENHA_BTG = "41589855876"


def _btg_workbook(path: Path, *, referencia="Junho/2026", vencimento="01/06",
                  compras=None, pagamentos=None, outros=20.0,
                  abas_extras=0) -> Path:
    """Monta uma planilha com o mesmo formato do extrato do BTG.

    Reproduz o que separa este arquivo dos outros e que a leitura precisa
    vencer: a coluna A vazia, as duas tabelas empilhadas em colunas diferentes,
    as células já tipadas (datetime e float, não texto pt-BR), a parcela dentro
    do nome do estabelecimento, o vencimento sem ano — e, de propósito, um
    "Outros valores" que entra no Total da Fatura sem virar lançamento nenhum.
    """
    compras = compras if compras is not None else [
        (datetime(2026, 3, 8), "Petz (3/3)", 132.00, "Parcela sem juros", "8134"),
        (datetime(2026, 4, 26), "Supermercado Confianca", 348.78, "Compra à vista", "8134"),
        (datetime(2026, 4, 29), "Steamgames", 40.02, "Compra internacional", "4108"),
        (datetime(2026, 5, 3), "Netflix", 44.90, "Compra à vista", "8134"),
    ]
    pagamentos = pagamentos if pagamentos is not None else [
        (datetime(2026, 5, 4), "Pagamento de fatura", -5104.93),
    ]

    internacionais = round(sum(v for _, _, v, tipo, _ in compras
                               if tipo == "Compra internacional"), 2)
    nacionais = round(sum(v for _, _, v, tipo, _ in compras
                          if tipo != "Compra internacional"), 2)
    pago = round(sum(v for _, _, v in pagamentos), 2)
    total = round(nacionais + internacionais + outros, 2)

    wb = Workbook()
    ws = wb.active
    ws.title = "Titular"
    grid: list[list] = [
        [None] * 8,
        [None] * 8,
        [None, "Fatura Cartão de Crédito", None, None, None, None, referencia, None],
        [None] * 8,
        [None] * 8,
        [None, "Fatura Atual", None, None, None, "Resumo", None, None],
        [None, "Período de Compras", None, "27/04 até 28/05", None,
         "Lançamentos Nacionais", None, nacionais],
        [None, "Vencimento", None, vencimento, None,
         "Lançamentos Internacionais", None, internacionais],
        [None, "Pagamento mínimo", None, 997.75, None, "Saque no crédito", None, 0.0],
        [None, "Pagamento parcial", None, 0.0, None, "Outros valores", None, outros],
        [None, None, None, None, None, "Total de créditos recebidos", None, 0.0],
        [None, None, None, None, None, "Juros e encargos", None, 0.0],
        [None, None, None, None, None, "Parcelamento de faturas", None, 0.0],
        [None, None, None, None, None, "Saldo fatura anterior e pagamentos", None, 0.0],
        [None, None, None, None, None, "Total da Fatura", None, total],
        [None] * 8,
        [None] * 8,
        [None, "Pagamentos feitos pelo cliente", None, None, pago, None, None, None],
        [None] * 8,
        [None, "Data", "Descrição", None, "Valor", None, None, None],
    ]
    grid += [[None, data, desc, None, valor, None, None, None]
             for data, desc, valor in pagamentos]
    grid += [
        [None] * 8,
        [None, "Total de compras e despesas", None, None, total, None, None, None],
        [None] * 8,
        [None, "Data", "Descrição", None, "Valor", "Tipo de compra",
         "Código de autorização", "Final Cartão"],
    ]
    grid += [[None, data, desc, None, valor, tipo, "OWJ4F1", cartao]
             for data, desc, valor, tipo, cartao in compras]

    for row in grid:
        ws.append(row)

    # Uma aba a mais COM tabela é o caso que o perfil recusa de propósito, em
    # vez de somar só a primeira e esconder as outras em silêncio.
    for i in range(abas_extras):
        extra = wb.create_sheet(f"Adicional {i + 1}")
        for row in [[None, "Data", "Descrição", None, "Valor", "Tipo de compra",
                     "Código de autorização", "Final Cartão"],
                    [None, datetime(2026, 5, 5), "Loja Do Outro", None, 10.0,
                     "Compra à vista", "ZZZZZZ", "9999"]]:
            extra.append(row)

    wb.save(path)
    return path


def _cifrar(origem: Path, destino: Path, senha: str) -> Path:
    """Grava `origem` cifrado com `senha` — é assim que o BTG manda a fatura.

    Cifrar DENTRO do teste, em vez de versionar um arquivo pronto, é o que
    mantém a suíte hermética e o repositório sem binário: não entra fatura de
    verdade num diretório de fixtures, e o arquivo é sempre reconstruído.
    """
    import msoffcrypto

    saida = io.BytesIO()
    with origem.open("rb") as fh:
        msoffcrypto.OfficeFile(fh).encrypt(senha, saida)
    destino.write_bytes(saida.getvalue())
    return destino


@pytest.fixture
def btg_xlsx(tmp_path) -> Path:
    return _btg_workbook(tmp_path / "btg-aberto.xlsx")


@pytest.fixture
def btg_xlsx_cifrado(tmp_path) -> Path:
    aberto = _btg_workbook(tmp_path / "btg-claro.xlsx")
    return _cifrar(aberto, tmp_path / "btg_extrato.xlsx", SENHA_BTG)


@pytest.fixture
def nubank_csv(tmp_path) -> Path:
    # O formato REAL do export do app, conferido contra
    # `documents/nubank_extrato_app.csv`: vírgula decimal entre aspas, e o sinal
    # de menos separado do número por um espaço. A versão anterior desta fixture
    # usava `270.51` — inventado, e o perfil tinha sido escrito para casar com a
    # invenção em vez de com o arquivo.
    path = tmp_path / "nubank.csv"
    path.write_text(
        "date,title,amount\n"
        '2026-07-03,Supermercados Alvorada,"270,51"\n'
        "2026-07-05,Uber *Trip,\"26,74\"\n"
        "2026-07-08,Amazon BR,\"59,13\"\n"
        "2026-07-11,Renner,\"79,96\"\n"
        '2026-07-12,Pagamento recebido,"- 1.000,00"\n',
        encoding="utf-8",
    )
    return path


# ---------------------------------------------------------------------------
# Config isolada por teste
# ---------------------------------------------------------------------------

@pytest.fixture
def config_dir(tmp_path) -> Path:
    """Cópia da config CONGELADA (tests/fixtures/config), gravável e isolada.

    Não é a configuração real de propósito: um teste que dependesse dela
    passaria a falhar assim que alguém confirmasse um chute ou mapeasse um
    estabelecimento novo pelo portal. Ver `test_config_real_e_valida` para a
    cobertura da configuração de verdade.
    """
    destino = tmp_path / "config"
    shutil.copytree(CONFIG, destino)
    return destino


@pytest.fixture
def client(config_dir, tmp_path, monkeypatch):
    monkeypatch.setenv("FATURA_RULES_PATH", str(config_dir / "categories.yml"))
    monkeypatch.setenv("FATURA_DB_PATH", str(tmp_path / "state.db"))

    # `get_settings` é cacheado com lru_cache; sem limpar, um teste herdaria a
    # config do anterior.
    from api import app as app_module
    from api.settings import get_settings

    get_settings.cache_clear()
    app_module._store = None

    from fastapi.testclient import TestClient
    with TestClient(app_module.app) as c:
        yield c

    get_settings.cache_clear()
    app_module._store = None


@pytest.fixture
def categories_text(config_dir) -> str:
    return (config_dir / "categories.yml").read_text(encoding="utf-8")


# ---------------------------------------------------------------------------
# CSV de saída montado pelo SCHEMA, nunca por nomes fixos
# ---------------------------------------------------------------------------
#
# Os nomes das colunas são configuráveis pelo portal: `Descrição` vira `Item`,
# `Valor (R$)` vira `Valor`. Todo teste que escrevia o cabeçalho na mão passou
# a falhar assim que essa configuração foi exercida de verdade — uma edição
# legítima no "Formato de saída" derrubou o build inteiro.
#
# Estas ajudas montam o CSV a partir do schema. O teste fala de PAPÉIS (data,
# categoria, descricao, valor, pago) e o NOME da coluna fica por conta da
# configuração de quem estiver rodando a suíte.

@pytest.fixture
def output_schema(config_dir):
    from core.profiles import ConfigSet
    return ConfigSet.load(config_dir).output


def csv_de_saida_texto(schema, linhas: list[dict], extras: tuple = ()) -> str:
    """Monta um CSV no formato de saída ATUAL.

    `linhas` usa os papéis como chave. `extras` acrescenta colunas que o portal
    não conhece (`Mês`, `Ano`), para provar que sobrevivem à ida e volta.
    """
    import csv as _csv
    import io as _io

    colunas = list(schema.colunas) + list(extras)
    buffer = _io.StringIO()
    writer = _csv.DictWriter(buffer, fieldnames=colunas, lineterminator="\n")
    writer.writeheader()
    for linha in linhas:
        row = schema.linha(
            data=linha.get("data", ""), categoria=linha.get("categoria", ""),
            descricao=linha.get("descricao", ""), valor=linha.get("valor", ""),
            pago=linha.get("pago", "x"),
        )
        row.update({c: linha.get(c, "") for c in extras})
        writer.writerow(row)
    return buffer.getvalue()


def com_um_chute(config_dir, categoria: str = "Casa", palavra: str = "CHUTOMETRO"):
    """Garante um `# ?` no categories.yml da cópia de teste.

    Os chutes são para o usuário confirmar ou apagar pela aba Regras — ou seja,
    o arquivo real tende a ZERO com o uso. Testes sobre chutes precisam plantar
    o seu, em vez de depender de o arquivo de alguém ainda ter algum.
    """
    alvo = config_dir / "categories.yml"
    texto = alvo.read_text(encoding="utf-8")
    marca = f"  {categoria}:\n"
    if marca not in texto:
        raise AssertionError(f"categoria {categoria} não existe no categories.yml")
    texto = texto.replace(marca, f"{marca}    - {palavra}    # ? plantado pelo teste\n", 1)
    alvo.write_text(texto, encoding="utf-8")
    return palavra


def com_uma_redundancia(config_dir, categoria: str = "Casa",
                        curta: str = "ZZTESTE", longa: str = "ZZTESTE LOJA"):
    """Planta um par redundante: a longa nunca vence, a curta já pega tudo.

    Mesmo motivo do `com_um_chute` — o arquivo real fica MAIS limpo com o uso,
    então depender das redundâncias que ele tinha ontem é garantir build
    vermelho amanhã.
    """
    alvo = config_dir / "categories.yml"
    texto = alvo.read_text(encoding="utf-8")
    marca = f"  {categoria}:\n"
    if marca not in texto:
        raise AssertionError(f"categoria {categoria} não existe no categories.yml")
    texto = texto.replace(marca, f"{marca}    - {curta}\n    - {longa}\n", 1)
    alvo.write_text(texto, encoding="utf-8")
    return curta, longa


def real_statements() -> list[Path]:
    return sorted(REAL_INPUT.glob("*.xls")) if REAL_INPUT.is_dir() else []
